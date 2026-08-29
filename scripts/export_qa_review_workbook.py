#!/usr/bin/env python3
"""Export an Excel workbook for the Human QA Manager to review the AI's output.

One row per call with everything the AI decided and the reason it gave, plus
empty HUMAN_* columns for the reviewer. The reviewer fills those in, sends the
file back, and scripts/import_qa_review_workbook.py reads it in.

Round-trip contract: `call_id` in column A is the key and must not be edited,
reordered or deleted. Everything else can be sorted or filtered freely.

Usage:
    python scripts/export_qa_review_workbook.py                    # all calls
    python scripts/export_qa_review_workbook.py --limit 300        # a sample
    python scripts/export_qa_review_workbook.py --out ~/Downloads/x.xlsx
"""
import argparse
import json
import os
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

load_dotenv()
BASE = Path(__file__).resolve().parent.parent
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HUMAN_FILL = PatternFill("solid", fgColor="C55A11")
AI_FILL = PatternFill("solid", fgColor="385723")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)


def sb(path, params, page=1000):
    """Page through PostgREST, which caps rows per request."""
    rows, offset = [], 0
    while True:
        h = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
             "Range": f"{offset}-{offset + page - 1}"}
        r = requests.get(f"{SUPABASE_URL}/rest/v1/{path}", headers=h, params=params, timeout=90)
        r.raise_for_status()
        batch = r.json()
        rows.extend(batch)
        if len(batch) < page:
            return rows
        offset += page


def dashboard_label(o, m, c):
    """Reproduce build_ci_dataset.py's score->label rule exactly.

    Closing third is weighted 1.4x and the neutral band is +/-0.15; both are
    documented in docs/sentiment_definition_for_review.md as unapproved
    defaults, so the reviewer sees the weighted number, not just the label.
    """
    if o is None and m is None and c is None:
        return None, None
    o, m, c = (o or 0.0), (m or 0.0), (c or 0.0)
    w = (o + m + c * 1.4) / 3.4
    return ("positive" if w > 0.15 else "negative" if w < -0.15 else "neutral"), round(w, 3)


def join(v, sep="; ", limit=400):
    if v is None:
        return ""
    if isinstance(v, list):
        out = sep.join(str(x) for x in v if x is not None)
    elif isinstance(v, dict):
        out = json.dumps(v, ensure_ascii=False)
    else:
        out = str(v)
    return out[:limit]


def style_header(ws, headers, human_from):
    for i, name in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=name)
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.fill = HUMAN_FILL if name.startswith("HUMAN_") else (
            AI_FILL if name.startswith("ai_") or name.startswith("qa_") else HEADER_FILL)
    # "B2", not ws.cell(row=2, column=2): referencing a cell object *creates* it,
    # which pushes max_row to 2 and makes the first append() land on row 3,
    # leaving a blank row that breaks the re-import key lookup.
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30


def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--out", default=str(Path.home() / "Downloads" /
                                         "Sunrooof - AI Call Review for QA.xlsx"))
    args = ap.parse_args()

    print("☁️  loading call summaries…")
    # call_id breaks the tie. sb() pages with Range:, and two calls sharing a
    # start_time leave Postgres free to order them differently per page, which
    # silently duplicates some rows and drops others.
    calls = sb("call_summaries", {"select": "*", "order": "start_time.desc,call_id.asc"})
    if args.limit:
        calls = calls[:args.limit]
    print(f"   {len(calls)} calls")

    print("📂 loading enrichment (sentiment journey, ASR confidence)…")
    enrich = {}
    for p in (BASE / "out" / "ci_enrichment").glob("*.json"):
        try:
            enrich[p.stem] = json.loads(p.read_text())
        except (OSError, ValueError):
            pass
    print(f"   {len(enrich)} enrichment files")

    wb = Workbook()

    # ── Sheet 1: Calls ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Calls"
    headers = [
        "call_id", "crm_link", "call_date", "agent", "customer", "duration_min", "language",
        "ai_summary",
        "ai_sentiment_zoho", "ai_sentiment_dashboard", "sent_opening", "sent_mid",
        "sent_closing", "sent_weighted", "ai_emotions", "ai_unresolved_negative",
        "ai_outcome", "ai_interest_level", "ai_conversion_likelihood", "ai_next_step_secured",
        "ai_politeness_5", "ai_professionalism_5", "ai_reason_professionalism",
        "ai_objections", "ai_red_flags", "ai_risk_flags", "ai_buying_signals",
        "ai_asr_confidence", "ai_talk_pct_agent",
        "qa_final_score", "qa_tier", "qa_auto_zero", "qa_critical_misses",
        "qa_red_flags", "qa_needs_review",
        "HUMAN_sentiment_verdict", "HUMAN_agrees_with_ai", "HUMAN_comments",
    ]
    style_header(ws, headers, None)

    for row in calls:
        cid = row["call_id"]
        e = enrich.get(cid, {})
        s = (e.get("sentiment") or {})
        label, weighted = dashboard_label(s.get("opening"), s.get("mid"), s.get("closing"))
        dur = row.get("duration_seconds") or 0
        ws.append([
            cid,
            f"https://crm.zoho.in/crm/tab/Calls/{cid}",
            (row.get("start_time") or "")[:10],
            row.get("agent"), row.get("customer"),
            round(dur / 60, 1), row.get("language"),
            join(row.get("summary"), limit=900),
            row.get("customer_sentiment"), label,
            s.get("opening"), s.get("mid"), s.get("closing"), weighted,
            join(s.get("emotions")), s.get("unresolved_negative"),
            row.get("call_outcome"), row.get("interest_level"),
            row.get("conversion_likelihood"), row.get("next_step_secured"),
            row.get("agent_politeness"), row.get("agent_professionalism"),
            join(row.get("professionalism_notes"), limit=500),
            join(row.get("objections")), join(row.get("red_flags")),
            join(row.get("risk_flags")), join(row.get("buying_signals")),
            (e.get("asr_confidence") if not isinstance(e.get("asr_confidence"), dict)
             else e["asr_confidence"].get("score")),
            row.get("agent_talk_pct"),
            row.get("qa_final_score"), row.get("qa_tier"), row.get("qa_auto_zero"),
            join(row.get("qa_critical_miss_codes")), join(row.get("qa_red_flag_codes")),
            row.get("qa_requires_human_review"),
            "", "", "",
        ])

    widths(ws, {"A": 20, "B": 16, "C": 11, "D": 18, "E": 20, "F": 9, "G": 13, "H": 60,
                "I": 15, "J": 17, "K": 9, "L": 9, "M": 9, "N": 11, "O": 22, "P": 12,
                "Q": 18, "R": 13, "S": 15, "T": 12, "U": 11, "V": 12, "W": 45,
                "X": 30, "Y": 26, "Z": 26, "AA": 30, "AB": 12, "AC": 11,
                "AD": 11, "AE": 12, "AF": 11, "AG": 16, "AH": 14, "AI": 12,
                "AJ": 20, "AK": 18, "AL": 55})
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=8).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=23).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r, column=38).alignment = Alignment(wrap_text=True, vertical="top")

    # Dropdowns keep the returned values clean enough to re-import without
    # guessing what "pos", "Positive " or "ok" were meant to mean.
    dv = DataValidation(type="list", formula1='"Positive,Neutral,Negative,Not Applicable"',
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"AJ2:AJ{ws.max_row}")
    dv2 = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv2)
    dv2.add(f"AK2:AK{ws.max_row}")
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: FAQs ─────────────────────────────────────────────────────
    faq_path = BASE / "out" / "faq_analysis.json"
    if faq_path.exists():
        fa = json.loads(faq_path.read_text())
        wf = wb.create_sheet("FAQs")
        fh = ["faq_id", "topic", "canonical_question", "total_asks",
              "answered_clearly", "answered_partially", "not_answered",
              "top_regions", "HUMAN_verdict", "HUMAN_comments"]
        style_header(wf, fh, None)
        for f in sorted(fa.get("faqs", []), key=lambda x: -(x.get("total_asks") or 0)):
            sc = f.get("status_counts") or {}
            reg = sorted((f.get("by_region") or {}).items(), key=lambda kv: -kv[1])[:3]
            wf.append([
                f.get("id"), f.get("topic"), f.get("canonical_question"),
                f.get("total_asks"), sc.get("answered_clearly"),
                sc.get("answered_partially"), sc.get("not_answered"),
                ", ".join(f"{k} ({v})" for k, v in reg),
                "", "",
            ])
        widths(wf, {"A": 8, "B": 20, "C": 75, "D": 11, "E": 15, "F": 17, "G": 13,
                    "H": 34, "I": 18, "J": 55})
        for r in range(2, wf.max_row + 1):
            wf.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        dv3 = DataValidation(type="list", allow_blank=True, showDropDown=False,
                             formula1='"Good FAQ,Reword,Merge with another,Drop"')
        wf.add_data_validation(dv3)
        dv3.add(f"I2:I{wf.max_row}")
        wf.auto_filter.ref = wf.dimensions

    # ── Sheet 3: How sentiment is decided ─────────────────────────────────
    wi = wb.create_sheet("How to use")
    wi["A1"] = "AI Call Review — for the Human QA Manager"
    wi["A1"].font = Font(bold=True, size=14)
    notes = [
        "",
        "WHAT TO DO",
        "  1. Review the AI's decision on each call in the 'Calls' sheet.",
        "  2. Fill the three orange HUMAN_ columns at the far right. Leave blank to skip a call.",
        "  3. On the 'FAQs' sheet, mark whether each canonical question is worded well.",
        "  4. Send the file back. It gets read straight back into the pipeline.",
        "",
        "DO NOT change or reorder column A (call_id) — it is how rows are matched on re-import.",
        "Sorting and filtering are fine; the call_id travels with its row.",
        "",
        "THE TWO SENTIMENT COLUMNS ARE DIFFERENT AND OFTEN DISAGREE",
        "  ai_sentiment_zoho      — one word, written into the Zoho call note. The model is",
        "                           given NO definition of positive/neutral/negative for this.",
        "  ai_sentiment_dashboard — what leadership sees on the dashboard. Computed by formula:",
        "                           weighted = (opening + mid + closing x 1.4) / 3.4",
        "                           positive if > +0.15, negative if < -0.15, else neutral.",
        "  sent_weighted          — that computed number, so you can see near-boundary cases.",
        "",
        "  Neither weighting nor the +/-0.15 band has been formally approved. About a third of",
        "  calls sit within 0.15 of a boundary, so those labels flip easily. Full detail is in",
        "  the 'Customer Sentiment Definition for Review' document.",
        "",
        "KNOWN ISSUES WORTH CHECKING AS YOU REVIEW",
        "  - Calls that never connected are labelled Neutral rather than Not Applicable.",
        "  - A 0.00 score usually means 'not assessed', not 'balanced'.",
        "  - Politeness tends to read as positive; a polite brush-off may be scored positive.",
        "  - Sentiment is judged from TEXT ONLY. The pipeline cannot hear tone, so sarcasm and",
        "    resigned agreement are not recoverable. Judge the words, as the AI must.",
        "",
        "THE qa_ COLUMNS (PSM scorecard)",
        "  qa_final_score / qa_tier come from the 100-point PSM scorecard. These are still",
        "  being generated and several scorecard rules are awaiting sign-off, so treat them as",
        "  provisional. qa_tier NOT_SCORED means the AI could not assess the call at all —",
        "  it is not a zero and must not be read as a failing grade.",
        "",
        "ai_asr_confidence is the transcription confidence. A low value means the transcript",
        "itself may be unreliable, so disagreements on those calls may be transcription, not AI.",
    ]
    for i, line in enumerate(notes, start=2):
        wi.cell(row=i, column=1, value=line)
        if line and not line.startswith("  ") and line.isupper():
            wi.cell(row=i, column=1).font = Font(bold=True)
    wi.column_dimensions["A"].width = 100

    out = Path(args.out).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)

    audited = sum(1 for c in calls if c.get("qa_final_score") is not None)
    print(f"\n✅ {out}")
    print(f"   Calls sheet : {len(calls)} rows  ({audited} with a PSM score so far)")
    if faq_path.exists():
        print(f"   FAQs sheet  : {len(json.loads(faq_path.read_text()).get('faqs', []))} canonical questions")
    print(f"   size        : {out.stat().st_size / 1024 / 1024:.1f} MB")


if __name__ == "__main__":
    main()
