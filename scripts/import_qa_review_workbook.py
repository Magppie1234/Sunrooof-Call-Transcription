#!/usr/bin/env python3
"""Read the QA Manager's filled-in workbook back into the pipeline.

Reads only the HUMAN_* columns, matched on call_id. Everything else in the file
is ignored, so it does not matter if the reviewer sorted, filtered or hid
columns — as long as call_id is intact.

Writes human verdicts to Supabase and reports agreement between the human and
each of the two AI sentiment fields, which is the number that answers "can this
replace a human QA".

Usage:
    python scripts/import_qa_review_workbook.py ~/Downloads/reviewed.xlsx
    python scripts/import_qa_review_workbook.py reviewed.xlsx --dry-run
"""
import argparse
import json
import os
from collections import Counter
from pathlib import Path

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()
BASE = Path(__file__).resolve().parent.parent
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
HEAD = {"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
        "Content-Type": "application/json"}

VALID = {"positive", "neutral", "negative", "not applicable"}


def norm(v):
    return str(v).strip().lower() if v is not None and str(v).strip() else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    path = Path(args.workbook).expanduser()
    if not path.exists():
        raise SystemExit(f"not found: {path}")

    wb = load_workbook(path, data_only=True)
    if "Calls" not in wb.sheetnames:
        raise SystemExit("no 'Calls' sheet — is this the review workbook?")
    ws = wb["Calls"]
    header = {c.value: i for i, c in enumerate(ws[1]) if c.value}

    for required in ("call_id", "HUMAN_sentiment_verdict", "HUMAN_comments"):
        if required not in header:
            raise SystemExit(f"missing column: {required}")

    reviewed, bad, rows = [], [], 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        cid = r[header["call_id"]]
        if not cid:
            continue
        rows += 1
        verdict = norm(r[header["HUMAN_sentiment_verdict"]])
        agrees = norm(r[header.get("HUMAN_agrees_with_ai", -1)]) if "HUMAN_agrees_with_ai" in header else None
        comment = r[header["HUMAN_comments"]]
        comment = str(comment).strip() if comment else None
        if not (verdict or agrees or comment):
            continue                      # untouched row
        if verdict and verdict not in VALID:
            bad.append((cid, verdict)); continue
        reviewed.append({
            "call_id": str(cid),
            "human_sentiment": verdict,
            "human_agrees_with_ai": {"yes": True, "no": False}.get(agrees),
            "human_comments": comment,
            "ai_sentiment_zoho": r[header.get("ai_sentiment_zoho", -1)] if "ai_sentiment_zoho" in header else None,
            "ai_sentiment_dashboard": r[header.get("ai_sentiment_dashboard", -1)] if "ai_sentiment_dashboard" in header else None,
        })

    print(f"📄 {rows} rows in the sheet")
    print(f"✍️  {len(reviewed)} reviewed by a human")
    if bad:
        print(f"⚠  {len(bad)} unrecognised verdict(s), skipped: {bad[:5]}")
    if not reviewed:
        print("nothing to import"); return

    # ── Agreement: the number that says whether the AI can stand in ──────
    for field in ("ai_sentiment_zoho", "ai_sentiment_dashboard"):
        pairs = [(x["human_sentiment"], norm(x[field])) for x in reviewed
                 if x["human_sentiment"] and x["human_sentiment"] != "not applicable"
                 and x.get(field)]
        if not pairs:
            continue
        agree = sum(1 for h, a in pairs if h == a)
        print(f"\n🤝 {field}: {agree}/{len(pairs)} agree ({agree / len(pairs) * 100:.0f}%)")
        mism = Counter(f"AI {a} -> human {h}" for h, a in pairs if h != a)
        for k, n in mism.most_common(5):
            print(f"     {n:3d}  {k}")

    na = sum(1 for x in reviewed if x["human_sentiment"] == "not applicable")
    if na:
        print(f"\n   {na} call(s) marked Not Applicable by the human "
              f"(currently counted as Neutral in the dashboard)")

    out = BASE / "out" / "human_qa_review.json"
    out.write_text(json.dumps(reviewed, indent=1, ensure_ascii=False))
    print(f"\n💾 {out}")

    if args.dry_run:
        print("dry run — nothing written to Supabase"); return

    ok = fail = 0
    for x in reviewed:
        body = {k: x[k] for k in ("human_sentiment", "human_agrees_with_ai", "human_comments")}
        r = requests.patch(f"{SUPABASE_URL}/rest/v1/call_summaries",
                           headers={**HEAD, "Prefer": "return=minimal"},
                           params={"call_id": f"eq.{x['call_id']}"},
                           data=json.dumps(body), timeout=30)
        if r.ok:
            ok += 1
        else:
            fail += 1
            if fail == 1:
                print(f"   ⚠ write failed: {r.status_code} {r.text[:200]}")
                if "human_sentiment" in r.text:
                    print("   → add the columns first:\n"
                          "     alter table call_summaries\n"
                          "       add column if not exists human_sentiment text,\n"
                          "       add column if not exists human_agrees_with_ai boolean,\n"
                          "       add column if not exists human_comments text;")
                    break
    print(f"✅ wrote {ok} verdict(s) to Supabase" + (f", {fail} failed" if fail else ""))

    # ── FAQ verdicts ─────────────────────────────────────────────────────
    if "FAQs" in wb.sheetnames:
        wf = wb["FAQs"]
        fh = {c.value: i for i, c in enumerate(wf[1]) if c.value}
        if "HUMAN_verdict" in fh:
            faq = []
            for r in wf.iter_rows(min_row=2, values_only=True):
                v = r[fh["HUMAN_verdict"]]
                c = r[fh.get("HUMAN_comments", -1)] if "HUMAN_comments" in fh else None
                if v or c:
                    faq.append({"faq_id": r[fh["faq_id"]],
                                "canonical_question": r[fh["canonical_question"]],
                                "verdict": v, "comments": c})
            if faq:
                p = BASE / "out" / "human_faq_review.json"
                p.write_text(json.dumps(faq, indent=1, ensure_ascii=False))
                print(f"✅ {len(faq)} FAQ verdict(s) -> {p}")


if __name__ == "__main__":
    main()
